import sqlite3
import os
import platform
import subprocess
import pandas as pd
import traceback
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout,
    QPushButton, QDateEdit, QSpinBox, QComboBox,
    QMessageBox, QListWidget, QListWidgetItem
)
from PyQt6.QtCore import QDate, Qt
from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QListWidgetItem, QInputDialog


class SinavProgramiWindow(QWidget):
    def __init__(self, bolum_id):
        self.istisna_sureleri = {}  # {ders_id: sure_dk}

        super().__init__()
        self.bolum_id = bolum_id
        self.setWindowTitle("📅 Sınav Programı Oluştur")
        self.resize(600, 600)
        self.setup_ui()
        self.load_dersler()

    # ---------------- UI ----------------
    def setup_ui(self):
        layout = QVBoxLayout()

        layout.addWidget(QLabel("Dahil Edilecek Dersler:"))
        self.ders_list = QListWidget()
        layout.addWidget(self.ders_list)

        tarih_layout = QHBoxLayout()
        tarih_layout.addWidget(QLabel("Başlangıç Tarihi:"))
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(QDate.currentDate())
        tarih_layout.addWidget(self.start_date)

        tarih_layout.addWidget(QLabel("Bitiş Tarihi:"))
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(QDate.currentDate().addDays(5))
        tarih_layout.addWidget(self.end_date)
        layout.addLayout(tarih_layout)

        tur_layout = QHBoxLayout()
        tur_layout.addWidget(QLabel("Sınav Türü:"))
        self.tur_combo = QComboBox()
        self.tur_combo.addItems(["Vize", "Final", "Bütünleme"])
        tur_layout.addWidget(self.tur_combo)
        layout.addLayout(tur_layout)

        sure_layout = QHBoxLayout()
        sure_layout.addWidget(QLabel("Sınav Süresi (dk):"))
        self.sure_spin = QSpinBox()
        self.sure_spin.setRange(30, 180)
        self.sure_spin.setValue(75)
        sure_layout.addWidget(self.sure_spin)

        sure_layout.addWidget(QLabel("Bekleme Süresi (dk):"))
        self.bekleme_spin = QSpinBox()
        self.bekleme_spin.setRange(5, 60)
        self.bekleme_spin.setValue(15)
        sure_layout.addWidget(self.bekleme_spin)
        layout.addLayout(sure_layout)


        # --- İstisnai süre butonu ---
        self.istisna_btn = QPushButton("🕒 İstisnai Sınav Süresi Belirle")
        self.istisna_btn.setToolTip("Bazı dersler için farklı sınav süresi tanımla")
        self.istisna_btn.clicked.connect(self.open_istisna_dialog)
        layout.addWidget(self.istisna_btn)


        self.olustur_btn = QPushButton("📅 Programı Oluştur")
        self.olustur_btn.clicked.connect(self.create_program)
        layout.addWidget(self.olustur_btn)

        self.setLayout(layout)

    # ---------------- Dersleri yükle ----------------
    def load_dersler(self):
        conn = sqlite3.connect("sinav_takvimi.db")
        cursor = conn.cursor()
        cursor.execute(
            "SELECT rowid AS id, ders_adi, ders_kodu, sinif FROM Dersler WHERE bolum_id=?",
            (self.bolum_id,)
        )
        dersler = cursor.fetchall()
        conn.close()

        self.ders_list.clear()
        for ders in dersler:
            item = QListWidgetItem(f"{ders[2]} - {ders[1]} (Sınıf: {ders[3]})")
            item.setData(Qt.ItemDataRole.UserRole, ders[0])
            item.setCheckState(Qt.CheckState.Checked)
            self.ders_list.addItem(item)



    from PyQt6.QtWidgets import QInputDialog

    def open_istisna_dialog(self):
        if self.ders_list.count() == 0:
            QMessageBox.warning(self, "Uyarı", "Henüz ders listesi yüklenmemiş.")
            return

        # 💡 Önce kullanıcıya bir ders seçtir
        dersler = []
        ders_map = {}
        for i in range(self.ders_list.count()):
            item = self.ders_list.item(i)
            ders_adi = item.text()
            ders_id = item.data(Qt.ItemDataRole.UserRole)
            dersler.append(ders_adi)
            ders_map[ders_adi] = ders_id

        ders_adi, ok = QInputDialog.getItem(
            self,
            "İstisnai Süre",
            "İstisna uygulanacak dersi seçin:",
            dersler,
            editable=False
        )
        if not ok or not ders_adi:
            return  # kullanıcı iptal ettiyse çık

        # 🕒 sonra sadece o ders için süre sor
        sure, ok = QInputDialog.getInt(
            self,
            "İstisnai Süre",
            f"{ders_adi} için sınav süresi (dakika):",
            75, 30, 180, 5
        )
        if not ok:
            return

        # 💾 kaydet
        ders_id = ders_map[ders_adi]
        self.istisna_sureleri[ders_id] = sure

        QMessageBox.information(
            self,
            "Bilgi",
            f"{ders_adi} için sınav süresi {sure} dk olarak kaydedildi."
        )



    # ---------------- Program oluşturma ----------------
    def create_program(self):
        from datetime import timedelta, datetime
        try:
            conn = sqlite3.connect("sinav_takvimi.db")
            cur = conn.cursor()

            # tablo yoksa oluştur
            cur.execute("""
                CREATE TABLE IF NOT EXISTS SinavProgrami (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ders_id INTEGER,
                    derslik_id INTEGER,
                    tarih TEXT,
                    saat TEXT,
                    sure INTEGER,
                    sinav_turu TEXT,
                    bolum_id INTEGER
                )
            """)
            conn.commit()

            baslangic = self.start_date.date().toPyDate()
            bitis = self.end_date.date().toPyDate()
            sinav_turu = self.tur_combo.currentText()
            sure = self.sure_spin.value()
            # 🔹 Her ders için özel süre varsa onu kullan
            

            min_gap = self.bekleme_spin.value()

            # seçilen dersler
            secili_dersler = [
                self.ders_list.item(i).data(Qt.ItemDataRole.UserRole)
                for i in range(self.ders_list.count())
                if self.ders_list.item(i).checkState() == Qt.CheckState.Checked
            ]
            if not secili_dersler:
                QMessageBox.warning(self, "Uyarı", "Lütfen en az bir ders seçin!")
                return

            # derslikler
            cur.execute(
                "SELECT rowid AS id, derslik_adi, kapasite, sira_sayisi, sutun_sayisi, duzen_tipi FROM Derslikler WHERE bolum_id=?",
                (self.bolum_id,)
            )
            derslikler = cur.fetchall()
            if not derslikler:
                QMessageBox.warning(self, "Uyarı", "Bu bölüme ait derslik bulunamadı!")
                return

            ders_bilgisi = {}
            ogrencisiz = []

            # ders bilgilerini al
            for ders_id in secili_dersler:
                ders_sure = self.istisna_sureleri.get(ders_id, sure)
                cur.execute("SELECT ders_kodu, ders_adi, sinif FROM Dersler WHERE rowid=?", (ders_id,))
                row = cur.fetchone()
                if not row:
                    continue
                ders_kodu, ders_adi, sinif = row
                ders_kodu = ders_kodu.strip().upper()

                cur.execute("""
                    SELECT DISTINCT ogrenci_no
                    FROM Ogrenci_Ders_Kayitlari
                    WHERE UPPER(ders_kodu)=? AND bolum_id=?
                """, (ders_kodu, self.bolum_id))
                ogrenciler = [r[0] for r in cur.fetchall()]

                if not ogrenciler:
                    ogrencisiz.append(f"{ders_kodu} - {ders_adi}")
                    continue

                ders_bilgisi[ders_id] = (ders_kodu, ders_adi, sinif, len(ogrenciler), ogrenciler)

            tarih_listesi = [
                baslangic + timedelta(days=i)
                for i in range((bitis - baslangic).days + 1)
                if (baslangic + timedelta(days=i)).weekday() < 5
            ]
            saat_listesi = ["09:00", "11:00", "13:00", "15:00"]

            cur.execute("DELETE FROM SinavProgrami WHERE bolum_id=?", (self.bolum_id,))
            conn.commit()

            ogrenci_program = {}
            slot_rooms = {}
            grade_next_day = {}
            ders_siralama = [d for d in secili_dersler if d in ders_bilgisi]
            ders_siralama.sort(key=lambda d: ders_bilgisi[d][3], reverse=True)
            program = []
            uyarilar = []
            doluluk_raporu = {}

            def uygun_derslik_bul(kac_ogr, day, hour):
                key = (day.isoformat(), hour)
                kullanilan = slot_rooms.get(key, set())
                uygun_odalar, yetersizler = [], []
                for (rid, rname, cap, sira, sutun, duzen) in derslikler:
                    if rid in kullanilan:
                        continue
                    duzen = duzen or "2'li"  # eğer None ise varsayılan olarak "2'li" kabul et
                    grup = 2 if "2" in duzen else 3

                    gercek_kapasite = sira * sutun * grup
                    if gercek_kapasite >= kac_ogr:
                        uygun_odalar.append((rid, rname, gercek_kapasite))
                    else:
                        yetersizler.append(f"{rname} ({gercek_kapasite}/{kac_ogr})")
                if uygun_odalar:
                    return min(uygun_odalar, key=lambda x: x[2])
                elif yetersizler:
                    raise ValueError("Kapasite yetersiz: " + ", ".join(yetersizler))
                return None

            def parse_time(hhmm):
                h, m = map(int, hhmm.split(":"))
                return h, m

            def slot_uygun_mu(ogrenciler, day, hour):
                h, m = parse_time(hour)
                slot_dt = datetime(day.year, day.month, day.day, h, m)
                for ogr in ogrenciler:
                    for var_dt in ogrenci_program.get(ogr, []):
                        diff = abs((slot_dt - var_dt).total_seconds()) / 60.0
                        if diff < min_gap:
                            return False
                return True

            for ders_id in ders_siralama:
                ders_kodu, ders_adi, sinif, ogr_say, ogr_list = ders_bilgisi[ders_id]
                yerlesti = False
                for gun in tarih_listesi:
                    for saat in saat_listesi:
                        if not slot_uygun_mu(ogr_list, gun, saat):
                            continue
                        try:
                            oda = uygun_derslik_bul(ogr_say, gun, saat)
                        except ValueError as ve:
                            uyarilar.append(f"⚠️ {ders_kodu} - {ders_adi}: {ve}")
                            continue
                        if not oda:
                            continue
                        yerlesti = True
                        cur.execute("""
                            INSERT INTO SinavProgrami (ders_id, derslik_id, tarih, saat, ders_sure, sinav_turu, bolum_id)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        """, (ders_id, oda[0], gun.strftime("%Y-%m-%d"), saat, sure, sinav_turu, self.bolum_id))
                        h, m = parse_time(saat)
                        slot_dt = datetime(gun.year, gun.month, gun.day, h, m)
                        for ogr in ogr_list:
                            ogrenci_program.setdefault(ogr, []).append(slot_dt)
                        slot_rooms.setdefault((gun.isoformat(), saat), set()).add(oda[0])
                        doluluk_raporu.setdefault(oda[1], []).append(ogr_say)
                        program.append([
                            f"{ders_kodu} - {ders_adi}",
                            oda[1],
                            gun.strftime("%d.%m.%Y"),
                            saat,
                            ders_sure,
                            sinav_turu
                        ])

                        break
                    if yerlesti:
                        break
                if not yerlesti:
                    uyarilar.append(f"⚠️ {ders_kodu} - {ders_adi}: Uygun slot veya derslik bulunamadı.")

            conn.commit()
            conn.close()

            if not program:
                QMessageBox.warning(self, "Bilgi", "Uygun program oluşturulamadı.")
                return

            file_path = os.path.abspath("sinav_programi.xlsx")
            df = pd.DataFrame(program, columns=["Ders", "Derslik", "Tarih", "Saat", "Süre (dk)", "Tür"])
            rapor = pd.DataFrame(
                [{"Derslik": d, "Ortalama Doluluk": f"{sum(lst)}/{len(lst)} sınav"} for d, lst in doluluk_raporu.items()]
            )
            uyaridf = pd.DataFrame(uyarilar + [f"Öğrencisiz ders: {x}" for x in ogrencisiz], columns=["Uyarılar"])

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Sınav Programı", index=False)
                rapor.to_excel(writer, sheet_name="Derslik Raporu", index=False)
                uyaridf.to_excel(writer, sheet_name="Uyarılar", index=False)

            wb = load_workbook(file_path)
            ws = wb["Sınav Programı"]
            renkler = ["#E3F2FD", "#E8F5E9", "#FFF3E0", "#F3E5F5"]
            row_color_map = {}
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3)):
                tarih = row[0].value
                if tarih not in row_color_map:
                    row_color_map[tarih] = renkler[len(row_color_map) % len(renkler)]
                renk = row_color_map[tarih]
                fill = PatternFill(start_color=renk.replace("#", ""), end_color=renk.replace("#", ""), fill_type="solid")
                for cell in ws[i + 2]:
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.font = Font(name="Segoe UI", size=11)
            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].width = 20
            wb.save(file_path)

            # otomatik aç
            system = platform.system()
            try:
                if system == "Windows":
                    os.startfile(file_path)
                elif system == "Darwin":
                    subprocess.call(["open", file_path])
                else:
                    subprocess.call(["xdg-open", file_path])
            except Exception:
                QMessageBox.warning(self, "Uyarı", "Excel otomatik açılamadı, dosya klasöre kaydedildi.")

            QMessageBox.information(self, "Başarılı", f"Sınav programı oluşturuldu ve açıldı:\n\n{file_path}")

        except Exception as e:
            detay = "".join(traceback.format_exception(type(e), e, e.__traceback__))
            QMessageBox.critical(self, "Hata", f"Sınav programı oluşturulurken hata oluştu:\n{detay}")
